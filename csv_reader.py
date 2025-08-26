import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side, Font
import matplotlib.pyplot as plt
import sys
filePath = str(sys.argv[1])
# filePath = 'network-test.csv'
OUTPUTFILE = filePath[:-4] + 'DATA.xlsx'
CPUREADINGSFILE = f'temp_{filePath[:-4]}.csv'
try:
    NETWORK_ENABLED = sys.argv[2]
except:
    NETWORK_ENABLED = False

# Read using multi-row headers
df = pd.read_csv(filePath, header=[2, 3])
# Show the top of the DataFrame
print(df.head())

df.to_excel(OUTPUTFILE, header=[2, 3])



# Load the workbook
wb = load_workbook(OUTPUTFILE)

# Select the active sheet (or use wb['SheetName'])
ws = wb.active


# Merge cells B2 to D2
ws.merge_cells('B4:D4')
ws.merge_cells('E4:H4')
ws.merge_cells('I4:P4')
ws.merge_cells('Q4:W4')
ws.merge_cells('X4:AA4')

if NETWORK_ENABLED:
    ws.merge_cells('AB4:AC4')
    ws.merge_cells('AD4:AE4')
    ws.merge_cells('AF4:AG4')
    ws.merge_cells('AH4:AI4')
    ws.merge_cells('AJ4:AK4')
    ws.merge_cells('AL4:AM4')

   

# Add CPU temp & freq readings
cf = pd.read_csv(CPUREADINGSFILE)
print(cf.head())

ws.merge_cells('AN4:AO4' if NETWORK_ENABLED else 'AB4:AC4')
ws.cell(row=4, column=40 if NETWORK_ENABLED else 28).value = 'CPU monitor'
ws.cell(row=5, column=40 if NETWORK_ENABLED else 28).value = 'Temperature'
ws.cell(row=5, column=41 if NETWORK_ENABLED else 29).value = 'Frequency'

temperatures = cf['Temperature']
frequencies = cf['Frequency']

# Load temperatures
for i, row in zip(range(0, len(temperatures)), range(6, ws.max_row + 1)):
    cell = ws.cell(row=row, column=40 if NETWORK_ENABLED else 28)
    cell.value = temperatures[i] / 1000 # Convert from milidegree C to C

# Load frequencies
for i, row in zip(range(0, len(frequencies)), range(6, ws.max_row + 1)):
    cell = ws.cell(row=row, column=41 if NETWORK_ENABLED else 29)
    cell.value = round(frequencies[i] / 1000000, 3) # Convert from raw kHz to GHz


# Choose the column (e.g., C is column 3)
for column in range(5, 17):
    for row in range(6, ws.max_row + 1):  # start at 2 to skip header
        cell = ws.cell(row=row, column=column)  # Column C
        if isinstance(float(cell.value), (int, float)):
            print("OLD VALUE: ", cell.value)
            cell.value = round(float(cell.value)/1024)
            print("NEW VALUE: ", cell.value)


# Convert percentage numbers from strings to numbers (floats)
for column in range(17, 27):
    for row in range(6, ws.max_row + 1):  # start at 2 to skip header
        cell = ws.cell(row=row, column=column)  # Column C
        if isinstance(cell.value, str):
            print("OLD VALUE: ", cell.value)
            cell.value = float(cell.value)
            print("NEW VALUE: ", cell.value)

for column in range(2, 5):
    for row in range(6, ws.max_row + 1):  # start at 2 to skip header
        cell = ws.cell(row=row, column=column)  # Column C
        if isinstance(cell.value, str):
            print("OLD VALUE: ", cell.value)
            cell.value = float(cell.value)
            print("NEW VALUE: ", cell.value)
            
# convert net/etho values to KB/s (this handles dstat -n)
for column in range(28, 33):
    for row in range(6, ws.max_row + 1):  # start at 2 to skip header
        cell = ws.cell(row=row, column=column)  # Column C
        if isinstance(cell.value, str):
            print("OLD VALUE: ", cell.value)
            cell.value = round(float(cell.value)/1024, 2)
            print("NEW VALUE: ", cell.value)
            
            
# convert pkt/eth0 to round number of packets received/sent
for column in range(34, 39):
    for row in range(6, ws.max_row + 1):  # start at 2 to skip header
        cell = ws.cell(row=row, column=column)  # Column C
        if isinstance(cell.value, str):
            print("OLD VALUE: ", cell.value)
            cell.value = round(float(cell.value))
            print("NEW VALUE: ", cell.value)


# convert timestamps into numbers
minute = -1
for row in range(6, ws.max_row + 1):  # start at 2 to skip header

    cell = ws.cell(row=row, column=1)  # Column A
    if isinstance(cell.value, str):
        print("OLD TIME: ", cell.value)
        cell.value = minute + 1
        print("NEW TIME: ", cell.value)
    minute += 1




# Define range (e.g., B2:D10)
for row in ws['A1:AC15']:
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')


# Define a thin border
thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

# Apply to each cell in the range
for row in ws['A4:AC15']: # I think this can be removed. The loop under is doing the full job
    for cell in row:
        cell.font = Font(bold=True)
        cell.border = border

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=41 if NETWORK_ENABLED else 29):
    for cell in row:
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Save changes to a new file or overwrite
wb.save(OUTPUTFILE)